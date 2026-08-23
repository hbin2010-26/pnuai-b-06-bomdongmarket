"""공간×작물 계산 결과를 Excel 통합문서로 저장한다."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_PATH = OUTPUT_DIR / "Profit_Output.xlsx"

NAVY = "0F172A"
EMERALD = "0F766E"
MINT = "D1FAE5"
SKY = "E0F2FE"
ROSE = "FFE4E6"
AMBER = "FEF3C7"
SLATE = "475569"
LIGHT_LINE = "DCE5EA"
WHITE = "FFFFFF"

CURRENCY_FORMAT = '#,##0;[Red](#,##0);-'
NUMBER_FORMAT = '#,##0.0;[Red](#,##0.0);-'
INTEGER_FORMAT = '#,##0;[Red](#,##0);-'
PERCENT_FORMAT = '0.0%'


def _section(site: dict[str, object], name: str) -> dict[str, Any]:
    value = site[name]
    if not isinstance(value, dict):
        raise TypeError(f"{name} 계산 결과가 dict 형식이 아닙니다.")
    return value


def _monthly_rows(site: dict[str, object], name: str) -> list[dict[str, Any]]:
    monthly = _section(site, name)["monthly"]
    if not isinstance(monthly, list):
        raise TypeError(f"{name}.monthly 계산 결과가 list 형식이 아닙니다.")
    return monthly


def _scenario_metadata(
    sites: list[dict[str, object]],
) -> tuple[list[str], list[str]]:
    """계산 결과의 공간과 작물 이름을 입력 순서대로 반환한다."""
    site_names = list(
        dict.fromkeys(
            str(_section(site, "space_row")["site_name"])
            for site in sites
        )
    )
    crop_names = list(
        dict.fromkeys(
            str(_section(site, "space_row")["crop_name"])
            for site in sites
        )
    )
    return site_names, crop_names


def _style_title(sheet: Any, cell_range: str, title: str) -> None:
    sheet.merge_cells(cell_range)
    cell = sheet[cell_range.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="맑은 고딕", size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[cell.row].height = 34


def _style_header(sheet: Any, row: int, start_col: int, end_col: int) -> None:
    for column in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = PatternFill("solid", fgColor=EMERALD)
        cell.font = Font(name="맑은 고딕", bold=True, color=WHITE)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            bottom=Side(style="medium", color="0B4F4A")
        )
    sheet.row_dimensions[row].height = 34


def _autofit_with_caps(
    sheet: Any,
    min_width: float = 10,
    max_width: float = 28,
) -> None:
    for column_cells in sheet.columns:
        column_index = column_cells[0].column
        longest = 0
        for cell in column_cells:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        width = min(max(longest + 2, min_width), max_width)
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def _add_table(sheet: Any, reference: str, name: str) -> None:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _create_summary_sheet(
    workbook: Workbook,
    sites: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("요약")
    sheet.sheet_view.showGridLines = False
    site_names, crop_names = _scenario_metadata(sites)
    comparison_label = f"{len(site_names)}×{len(crop_names)}"

    headers = [
        "시나리오ID",
        "사업장ID",
        "사업장명",
        "작물",
        "공실 전체면적(m²)",
        "사용가능 바닥면적(m²)",
        "재배면적(m²)",
        "원하는 월세(원)",
        "월 총생산량(kg)",
        "월 판매량(kg)",
        "월 매출(원)",
        "월평균 전력량(kWh)",
        "월 전기비(원)",
        "배액률",
        "월 작물 순소비량(m³)",
        "월 배액량(m³)",
        "월 기타 용수량(m³)",
        "월 총 용수량(m³)",
        "월 수도비(원)",
        "월 재료비(원)",
        "월 인건비(원)",
        "월 기기 대여비(원)",
        "기타비용(원)",
        "월 운영비(원)",
        "월 영업이익(원)",
        "공간 대여자 배분율",
        "공간 대여자 예상수익(원)",
        "공실 월세 대비 차이(원)",
        "사업장 영업이익(원)",
        "추천 방식",
        "계약 형태",
    ]
    _style_title(
        sheet,
        "A1:AE1",
        f"Profit Calculator 1.0.1 · {comparison_label} 수익성 비교",
    )
    sheet["A2"] = "생성 시각"
    sheet["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet["D2"] = "분석 기준"
    sheet["E2"] = (
        f"{len(site_names)}개 공간 × {len(crop_names)}개 작물"
        f" = {len(sites)}개 대안 시나리오"
    )
    sheet["A2"].font = sheet["D2"].font = Font(
        name="맑은 고딕", bold=True, color=SLATE
    )
    sheet.append([])
    sheet.append(headers)
    _style_header(sheet, 4, 1, len(headers))

    for site in sites:
        meta = _section(site, "space_row")
        space = _section(site, "space")
        production = _section(site, "production")
        sales = _section(site, "sales")
        electricity = _section(site, "electricity")
        water = _section(site, "water")
        material = _section(site, "material")
        labor = _section(site, "labor")
        profit = _section(site, "profit")
        sheet.append(
            [
                site["scenario_id"],
                meta["site_id"],
                meta["site_name"],
                meta["crop_name"],
                space["total_area_m2"],
                space["available_floor_area_m2"],
                space["cultivation_area_m2"],
                profit["desired_monthly_rent_krw"],
                production["monthly_total_production_kg"],
                production["monthly_sales_kg"],
                sales["monthly_revenue_krw"],
                electricity["average_monthly_energy_kwh"],
                electricity["monthly_electricity_cost_krw"],
                water["drainage_ratio"],
                float(water["monthly_evapotranspiration_l"]) / 1000.0,
                float(water["monthly_drainage_l"]) / 1000.0,
                float(water["monthly_other_water_l"]) / 1000.0,
                water["monthly_total_water_m3"],
                water["monthly_water_cost_krw"],
                material["monthly_material_cost_krw"],
                labor["monthly_labor_cost_krw"],
                profit["monthly_equipment_rental_cost_krw"],
                profit["monthly_other_cost_krw"],
                profit["monthly_operating_cost_krw"],
                profit["monthly_operating_profit_krw"],
                profit["landlord_share_ratio"],
                profit["landlord_expected_income_krw"],
                profit["rent_income_difference_krw"],
                profit["business_operating_profit_krw"],
                profit["recommendation"],
                profit["contract_type"],
            ]
        )

    last_row = 4 + len(sites)
    _add_table(sheet, f"A4:AE{last_row}", "ProfitScenarioSummary")
    sheet.freeze_panes = "D5"
    sheet.auto_filter.ref = f"A4:AE{last_row}"

    for row in range(5, last_row + 1):
        for column in range(1, 32):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="center")
        for column in (5, 6, 7, 9, 10, 15, 16, 17, 18):
            sheet.cell(row=row, column=column).number_format = NUMBER_FORMAT
        for column in (8, 11, 13, 19, 20, 21, 22, 23, 24, 25, 27, 28, 29):
            sheet.cell(row=row, column=column).number_format = CURRENCY_FORMAT
        sheet.cell(row=row, column=12).number_format = INTEGER_FORMAT
        sheet.cell(row=row, column=14).number_format = PERCENT_FORMAT
        sheet.cell(row=row, column=26).number_format = PERCENT_FORMAT

    negative_fill = PatternFill("solid", fgColor=ROSE)
    positive_fill = PatternFill("solid", fgColor=MINT)
    for column in ("Y", "AA", "AB", "AC"):
        target = f"{column}5:{column}{last_row}"
        sheet.conditional_formatting.add(
            target,
            CellIsRule(operator="lessThan", formula=["0"], fill=negative_fill),
        )
    sheet.conditional_formatting.add(
        f"AE5:AE{last_row}",
        FormulaRule(formula=["$AE5=\"장기계약형\""], fill=positive_fill),
    )
    sheet.conditional_formatting.add(
        f"AE5:AE{last_row}",
        FormulaRule(formula=["$AE5=\"단기계약형\""], fill=PatternFill("solid", fgColor=AMBER)),
    )

    _autofit_with_caps(sheet, max_width=26)
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["AD"].width = 34
    sheet.column_dimensions["AE"].width = 15


def _create_cost_breakdown_sheet(
    workbook: Workbook,
    sites: list[dict[str, object]],
) -> None:
    """각 시나리오의 월 운영비 구성금액과 구성비를 표시한다."""
    sheet = workbook.create_sheet("비용구성")
    sheet.sheet_view.showGridLines = False
    first_profit = _section(sites[0], "profit")
    equipment_rental_rate = float(
        first_profit["equipment_rental_rate_krw_m2_month"]
    )
    _style_title(sheet, "A1:Q1", "시나리오별 월 비용 구성 및 운영비 대비 비율")
    sheet["A2"] = "비율 기준"
    sheet["B2"] = "각 비용 ÷ 월 운영비"
    sheet["D2"] = "기기 대여비 기준"
    sheet["E2"] = (
        f"사용가능 바닥면적 × {equipment_rental_rate:,.0f}원/m²/month"
    )
    sheet["A2"].font = sheet["D2"].font = Font(
        name="맑은 고딕", bold=True, color=SLATE
    )
    headers = [
        "시나리오ID",
        "사업장",
        "작물",
        "월 운영비(원)",
        "전기비(원)",
        "전기비 비율",
        "수도비(원)",
        "수도비 비율",
        "재료비(원)",
        "재료비 비율",
        "인건비(원)",
        "인건비 비율",
        "기기 대여비(원)",
        "기기 대여비 비율",
        "기타비용(원)",
        "기타비용 비율",
        "비용 비율 합계",
    ]
    sheet.append([])
    sheet.append(headers)
    _style_header(sheet, 4, 1, len(headers))

    for site in sites:
        meta = _section(site, "space_row")
        electricity = _section(site, "electricity")
        water = _section(site, "water")
        material = _section(site, "material")
        labor = _section(site, "labor")
        profit = _section(site, "profit")
        row = sheet.max_row + 1
        sheet.append(
            [
                site["scenario_id"],
                meta["site_name"],
                meta["crop_name"],
                profit["monthly_operating_cost_krw"],
                electricity["monthly_electricity_cost_krw"],
                f"=IF($D{row}=0,0,E{row}/$D{row})",
                water["monthly_water_cost_krw"],
                f"=IF($D{row}=0,0,G{row}/$D{row})",
                material["monthly_material_cost_krw"],
                f"=IF($D{row}=0,0,I{row}/$D{row})",
                labor["monthly_labor_cost_krw"],
                f"=IF($D{row}=0,0,K{row}/$D{row})",
                profit["monthly_equipment_rental_cost_krw"],
                f"=IF($D{row}=0,0,M{row}/$D{row})",
                profit["monthly_other_cost_krw"],
                f"=IF($D{row}=0,0,O{row}/$D{row})",
                f"=SUM(F{row},H{row},J{row},L{row},N{row},P{row})",
            ]
        )

    last_row = 4 + len(sites)
    _add_table(sheet, f"A4:Q{last_row}", "ProfitCostBreakdown")
    sheet.freeze_panes = "D5"
    for row in range(5, last_row + 1):
        for column in (4, 5, 7, 9, 11, 13, 15):
            sheet.cell(row=row, column=column).number_format = CURRENCY_FORMAT
        for column in (6, 8, 10, 12, 14, 16, 17):
            sheet.cell(row=row, column=column).number_format = PERCENT_FORMAT
            sheet.cell(row=row, column=column).fill = PatternFill(
                "solid", fgColor=SKY
            )
    sheet.conditional_formatting.add(
        f"Q5:Q{last_row}",
        FormulaRule(
            formula=["ABS($Q5-1)>0.000001"],
            fill=PatternFill("solid", fgColor=ROSE),
        ),
    )
    _autofit_with_caps(sheet, max_width=24)
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["O"].width = 24


def _detail_metrics(site: dict[str, object]) -> list[tuple[str, str, float, str]]:
    space = _section(site, "space")
    production = _section(site, "production")
    sales = _section(site, "sales")
    hvac = _section(site, "hvac")
    humidity = _section(site, "humidity")
    electricity = _section(site, "electricity")
    water = _section(site, "water")
    material = _section(site, "material")
    labor = _section(site, "labor")
    profit = _section(site, "profit")
    return [
        ("1. 공간", "공실 전체면적", float(space["total_area_m2"]), "m²"),
        ("1. 공간", "사용가능 바닥면적", float(space["available_floor_area_m2"]), "m²"),
        ("1. 공간", "작물별 모듈 층 수", float(space["module_layers"]), "층"),
        ("1. 공간", "재배면적", float(space["cultivation_area_m2"]), "m²"),
        ("1. 공간", "공간 체적", float(space["volume_m3"]), "m³"),
        ("2. 생산", "월 총생산량", float(production["monthly_total_production_kg"]), "kg"),
        ("2. 생산", "월 판매량", float(production["monthly_sales_kg"]), "kg"),
        ("3. 매출", "판매가격", float(sales["price_krw_kg"]), "원/kg"),
        ("3. 매출", "월 매출", float(sales["monthly_revenue_krw"]), "원"),
        ("4. 조명·냉난방", "필요 조명전력", float(hvac["lighting_power_w"]), "W"),
        ("4. 조명·냉난방", "월 조명전력량", float(hvac["lighting_energy_kwh_month"]), "kWh"),
        ("5. 습도", "월 증발산량", float(humidity["monthly_evapotranspiration_kg"]), "kg"),
        ("6. 전기비", "월평균 총 전력량", float(electricity["average_monthly_energy_kwh"]), "kWh"),
        ("6. 전기비", "월 전기비", float(electricity["monthly_electricity_cost_krw"]), "원"),
        ("7. 수도비", "배액률", float(water["drainage_ratio"]), "%"),
        ("7. 수도비", "월 작물 순소비량", float(water["monthly_evapotranspiration_l"]) / 1000.0, "m³"),
        ("7. 수도비", "월 배액량", float(water["monthly_drainage_l"]) / 1000.0, "m³"),
        ("7. 수도비", "월 작물 관수량", float(water["monthly_crop_irrigation_l"]) / 1000.0, "m³"),
        ("7. 수도비", "월 기타 용수량", float(water["monthly_other_water_l"]) / 1000.0, "m³"),
        ("7. 수도비", "월 총 용수량", float(water["monthly_total_water_m3"]), "m³"),
        ("7. 수도비", "월 수도비", float(water["monthly_water_cost_krw"]), "원"),
        (
            "8. 재료비",
            "면적당 월 환산 모종비",
            float(material["seedling_cost_per_m2_month_krw"]),
            "원/m²/month",
        ),
        ("8. 재료비", "월 모종비", float(material["monthly_seedling_cost_krw"]), "원"),
        ("8. 재료비", "월 양액비", float(material["monthly_nutrient_cost_krw"]), "원"),
        ("8. 재료비", "월 재료비", float(material["monthly_material_cost_krw"]), "원"),
        ("9. 인건비", "월 노동시간", float(labor["monthly_labor_hours"]), "시간"),
        ("9. 인건비", "월 인건비", float(labor["monthly_labor_cost_krw"]), "원"),
        ("10. 수익", "기기 대여비 적용면적", float(profit["equipment_rental_area_m2"]), "m²"),
        ("10. 수익", "면적당 월 기기 대여비", float(profit["equipment_rental_rate_krw_m2_month"]), "원/m²/month"),
        ("10. 수익", "월 기기 대여비", float(profit["monthly_equipment_rental_cost_krw"]), "원"),
        ("10. 수익", "월 운영비", float(profit["monthly_operating_cost_krw"]), "원"),
        ("10. 수익", "월 영업이익", float(profit["monthly_operating_profit_krw"]), "원"),
        ("10. 수익", "공간 대여자 예상수익", float(profit["landlord_expected_income_krw"]), "원"),
        ("10. 수익", "원하는 월세", float(profit["desired_monthly_rent_krw"]), "원"),
        ("10. 수익", "공실 월세 대비 차이", float(profit["rent_income_difference_krw"]), "원"),
        ("10. 수익", "사업장 영업이익", float(profit["business_operating_profit_krw"]), "원"),
    ]


def _create_detail_sheet(
    workbook: Workbook,
    sites: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("계산상세")
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, "A1:G1", "계산 블록 1~10 상세 결과")
    headers = ["시나리오ID", "사업장", "작물", "계산 블록", "항목", "값", "단위"]
    sheet.append([])
    sheet.append(headers)
    _style_header(sheet, 3, 1, len(headers))

    for site in sites:
        meta = _section(site, "space_row")
        for block, item, value, unit in _detail_metrics(site):
            sheet.append(
                [
                    site["scenario_id"],
                    meta["site_name"],
                    meta["crop_name"],
                    block,
                    item,
                    value,
                    unit,
                ]
            )
    last_row = sheet.max_row
    _add_table(sheet, f"A3:G{last_row}", "ProfitCalculationDetail")
    sheet.freeze_panes = "D4"
    for row in range(4, last_row + 1):
        unit = sheet.cell(row=row, column=7).value
        if unit in {"원", "원/kg", "원/m²/month"}:
            sheet.cell(row=row, column=6).number_format = CURRENCY_FORMAT
        elif unit == "%":
            sheet.cell(row=row, column=6).number_format = PERCENT_FORMAT
        elif unit == "층":
            sheet.cell(row=row, column=6).number_format = INTEGER_FORMAT
        else:
            sheet.cell(row=row, column=6).number_format = NUMBER_FORMAT
    _autofit_with_caps(sheet)
    sheet.column_dimensions["E"].width = 26


def _create_energy_sheet(
    workbook: Workbook,
    sites: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("월별전력량")
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, "A1:L1", "12개월 환경제어 전력량")
    headers = [
        "시나리오ID",
        "사업장",
        "작물",
        "월",
        "외기온도(°C)",
        "외기상대습도",
        "조명(kWh)",
        "난방(kWh)",
        "냉방(kWh)",
        "제습(kWh)",
        "가습(kWh)",
        "총합(kWh)",
    ]
    sheet.append([])
    sheet.append(headers)
    _style_header(sheet, 3, 1, len(headers))

    for site in sites:
        meta = _section(site, "space_row")
        hvac_monthly = _monthly_rows(site, "hvac")
        electricity_monthly = _monthly_rows(site, "electricity")
        for hvac_row, energy_row in zip(
            hvac_monthly, electricity_monthly, strict=True
        ):
            sheet.append(
                [
                    site["scenario_id"],
                    meta["site_name"],
                    meta["crop_name"],
                    energy_row["month"],
                    hvac_row["outdoor_temperature_c"],
                    hvac_row["outdoor_relative_humidity"],
                    energy_row["lighting_energy_kwh"],
                    energy_row["heating_energy_kwh"],
                    energy_row["cooling_energy_kwh"],
                    energy_row["dehumidification_energy_kwh"],
                    energy_row["humidification_energy_kwh"],
                    energy_row["total_environment_energy_kwh"],
                ]
            )
    last_row = sheet.max_row
    _add_table(sheet, f"A3:L{last_row}", "MonthlyEnergyDetail")
    sheet.freeze_panes = "D4"
    for row in range(4, last_row + 1):
        sheet.cell(row=row, column=5).number_format = "0.0"
        sheet.cell(row=row, column=6).number_format = "0.00"
        for column in range(7, 13):
            sheet.cell(row=row, column=column).number_format = INTEGER_FORMAT
    _autofit_with_caps(sheet)


def _create_assumptions_sheet(
    workbook: Workbook,
    sites: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("입력기준")
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, "A1:D1", "모델 기준 및 입력 파일")
    headers = ["구분", "항목", "값", "설명"]
    sheet.append([])
    sheet.append(headers)
    _style_header(sheet, 3, 1, len(headers))
    water = _section(sites[0], "water")
    material = _section(sites[0], "material")
    profit = _section(sites[0], "profit")
    equipment_rental_rate = float(
        profit["equipment_rental_rate_krw_m2_month"]
    )
    _, crop_names = _scenario_metadata(sites)
    cultivable_ratios = list(
        dict.fromkeys(
            float(_section(site, "space_row")["cultivable_ratio"])
            for site in sites
        )
    )
    cultivable_ratio_summary = " · ".join(
        f"{ratio:.2f}" for ratio in cultivable_ratios
    )
    layers_by_crop: dict[str, float] = {}
    for site in sites:
        meta = _section(site, "space_row")
        space = _section(site, "space")
        layers_by_crop.setdefault(
            str(meta["crop_name"]), float(space["module_layers"])
        )
    layer_summary = " · ".join(
        f"{crop_name} {layers:g}층"
        for crop_name, layers in layers_by_crop.items()
    )
    rows = [
        ("버전", "프로그램 버전", "1.0.1", "6개 작물 비교 및 현실화한 비용 산정 반영"),
        ("시나리오", "조합 방식", "공간 × 작물", "모든 공간에 모든 작물을 각각 적용"),
        (
            "공간",
            "재배가능 비율",
            cultivable_ratio_summary,
            "space_info.csv 입력; 현재 모든 공간에 동일 적용",
        ),
        (
            "비용",
            "면적당 월 기기 대여비",
            f"{equipment_rental_rate:,.0f} 원/m²/month",
            "사용가능 바닥면적 기준",
        ),
        (
            "비용",
            "월 기타비용",
            f"{float(profit['monthly_other_cost_krw']):,.0f} 원/month",
            "standard_info.csv 공통값; 모든 시나리오에 월 1회 적용",
        ),
        ("수도", "배액률", f"{float(water['drainage_ratio']):.0%}", "standard_info.csv 공통값; 모든 작물에 동일 적용"),
        ("수도", "수도 종합단가", f"{float(water['water_rate_krw_m3']):,.0f} 원/m³", "standard_info.csv 입력"),
        ("수익", "공간 대여자 배분비율", "0.8", "contraction_info.csv 입력"),
        (
            "재료비",
            "모종비 적용 기준",
            "기존 1회 단가의 1/3을 매월 적용",
            "기존 단가로 연 4회 구입하는 것과 같은 연간 비용",
        ),
        (
            "재료비",
            "양액 단가",
            f"{float(material['nutrient_cost_per_l_krw']):,.0f} 원/L",
            "standard_info.csv 입력; 순환식 손실·보충 보정계수 1.1 적용",
        ),
        ("추천", "장기계약형", "예상수익 ≥ 원하는 월세", "두 금액이 같아도 장기계약형"),
        ("추천", "단기계약형", "적자 또는 예상수익 < 원하는 월세", "적자 금액은 그대로 출력"),
        ("입력 파일", "공간정보", "data/space_info.csv", "사업장별 원하는 월세 포함"),
        (
            "입력 파일",
            "작물 생산정보",
            "data/crop_production_info.csv",
            f"{'·'.join(crop_names)} 및 작물별 모듈 층 수",
        ),
        ("작물", "작물별 모듈 층 수", layer_summary, "crop_production_info.csv 입력"),
        ("입력 파일", "작물 판매정보", "data/crop_sale_info.csv", "작물별 kg당 판매가격"),
        (
            "주의",
            f"{len(sites)}개 결과 합산",
            "비교용",
            "같은 공간의 대안 작물을 동시에 운영한 합계가 아님",
        ),
    ]
    for row in rows:
        sheet.append(row)
    for row_index in range(4, sheet.max_row + 1):
        for column_index in range(1, 5):
            sheet.cell(row_index, column_index).alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )
    sheet.row_dimensions[14].height = 38
    sheet.row_dimensions[15].height = 38
    _add_table(sheet, f"A3:D{sheet.max_row}", "ProfitInputGuide")
    sheet.freeze_panes = "A4"
    _autofit_with_caps(sheet, max_width=42)
    sheet.column_dimensions["D"].width = 44


def _create_checks_sheet(
    workbook: Workbook,
    sites: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("검증")
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, "A1:G1", "계산 결과 검증")
    sheet["A2"] = "MODEL STATUS"
    sheet["A2"].font = Font(name="맑은 고딕", bold=True, color=SLATE)
    sheet["B2"] = f'=IF(COUNTIF(G5:G{4 + len(sites) * 3},"FAIL")=0,"PASS","FAIL")'
    sheet["B2"].font = Font(name="맑은 고딕", bold=True)
    headers = ["시나리오ID", "검증 항목", "실제값", "기대값", "차이", "허용오차", "상태"]
    sheet.append([])
    sheet.append(headers)
    _style_header(sheet, 4, 1, len(headers))

    for offset, site in enumerate(sites):
        summary_row = 5 + offset
        cost_row = 5 + offset
        profit_row = 5 + offset * 3
        water_row = profit_row + 1
        cost_check_row = profit_row + 2

        sheet.cell(profit_row, 1, site["scenario_id"])
        sheet.cell(profit_row, 2, "월 영업이익 = 공간 대여자 예상수익 + 사업장 영업이익")
        sheet.cell(profit_row, 3, f"='요약'!Y{summary_row}")
        sheet.cell(profit_row, 4, f"='요약'!AA{summary_row}+'요약'!AC{summary_row}")
        sheet.cell(profit_row, 5, f"=C{profit_row}-D{profit_row}")
        sheet.cell(profit_row, 6, 0.5)
        sheet.cell(profit_row, 7, f'=IF(ABS(E{profit_row})<=F{profit_row},"PASS","FAIL")')
        for column in range(3, 7):
            sheet.cell(profit_row, column).number_format = CURRENCY_FORMAT

        sheet.cell(water_row, 1, site["scenario_id"])
        sheet.cell(water_row, 2, "월 총 용수량 = 작물 순소비량 + 배액량 + 기타 용수량")
        sheet.cell(water_row, 3, f"='요약'!R{summary_row}")
        sheet.cell(water_row, 4, f"='요약'!O{summary_row}+'요약'!P{summary_row}+'요약'!Q{summary_row}")
        sheet.cell(water_row, 5, f"=C{water_row}-D{water_row}")
        sheet.cell(water_row, 6, 0.001)
        sheet.cell(water_row, 7, f'=IF(ABS(E{water_row})<=F{water_row},"PASS","FAIL")')
        for column in range(3, 7):
            sheet.cell(water_row, column).number_format = "0.000"

        sheet.cell(cost_check_row, 1, site["scenario_id"])
        sheet.cell(cost_check_row, 2, "월 운영비 = 전기비 + 수도비 + 재료비 + 인건비 + 기기 대여비 + 기타비용")
        sheet.cell(cost_check_row, 3, f"='비용구성'!D{cost_row}")
        sheet.cell(
            cost_check_row,
            4,
            f"='비용구성'!E{cost_row}+'비용구성'!G{cost_row}"
            f"+'비용구성'!I{cost_row}+'비용구성'!K{cost_row}"
            f"+'비용구성'!M{cost_row}+'비용구성'!O{cost_row}",
        )
        sheet.cell(cost_check_row, 5, f"=C{cost_check_row}-D{cost_check_row}")
        sheet.cell(cost_check_row, 6, 0.5)
        sheet.cell(
            cost_check_row,
            7,
            f'=IF(ABS(E{cost_check_row})<=F{cost_check_row},"PASS","FAIL")',
        )
        for column in range(3, 7):
            sheet.cell(cost_check_row, column).number_format = CURRENCY_FORMAT

    last_row = 4 + len(sites) * 3
    _add_table(sheet, f"A4:G{last_row}", "ProfitModelChecks")
    sheet.conditional_formatting.add(
        f"G5:G{last_row}",
        FormulaRule(formula=["$G5=\"PASS\""], fill=PatternFill("solid", fgColor=MINT)),
    )
    sheet.conditional_formatting.add(
        f"G5:G{last_row}",
        FormulaRule(formula=["$G5=\"FAIL\""], fill=PatternFill("solid", fgColor=ROSE)),
    )
    sheet.freeze_panes = "A5"
    _autofit_with_caps(sheet, max_width=50)
    sheet.column_dimensions["B"].width = 72


def write_profit_output(
    sites: list[dict[str, object]],
    output_path: Path | None = None,
) -> Path:
    """계산 결과를 ``output/Profit_Output.xlsx``에 저장하고 경로를 반환한다."""
    if not sites:
        raise ValueError("Excel로 저장할 계산 결과가 없습니다.")

    target = output_path or OUTPUT_PATH
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    _create_summary_sheet(workbook, sites)
    _create_cost_breakdown_sheet(workbook, sites)
    _create_detail_sheet(workbook, sites)
    _create_energy_sheet(workbook, sites)
    _create_assumptions_sheet(workbook, sites)
    _create_checks_sheet(workbook, sites)
    workbook.active = 0
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    try:
        workbook.save(target)
    except PermissionError as error:
        raise PermissionError(
            f"{target.name} 파일이 Excel에서 열려 있습니다. 파일을 닫고 다시 실행해 주세요."
        ) from error
    return target
