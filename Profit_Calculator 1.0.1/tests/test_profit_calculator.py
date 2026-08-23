"""1.0.1 핵심 계산과 Excel 출력 회귀 테스트."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from excel_output import write_profit_output
from main import calculate_all_sites, read_csv_rows
from material_cost_calculation import calculate_material_cost
from profit_calculation import (
    LONG_TERM_RECOMMENDATION,
    SHORT_TERM_RECOMMENDATION,
    calculate_profit,
)
from water_cost_calculation import calculate_water_cost


class ProfitCalculationTest(unittest.TestCase):
    def test_optional_web_interface_and_dependencies_are_removed(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        requirements = (project_dir / "requirements.txt").read_text(
            encoding="utf-8"
        ).lower()

        self.assertFalse((project_dir / "app.py").exists())
        self.assertEqual(requirements.strip(), "openpyxl>=3.1")

    def test_water_cost_includes_drainage_ratio(self) -> None:
        result = calculate_water_cost(
            {"total_area_m2": 100.0},
            {"monthly_evapotranspiration_kg": 1_000.0},
            {
                "drainage_ratio": 0.3,
                "other_water_l_m2_day": 0.2,
                "water_rate_krw_m3": 2_300.0,
            },
        )

        expected_crop_irrigation_l = 1_000.0 / 0.7
        expected_drainage_l = expected_crop_irrigation_l - 1_000.0
        expected_other_water_l = 100.0 * 0.2 * (365.0 / 12.0)
        expected_total_m3 = (
            expected_crop_irrigation_l + expected_other_water_l
        ) / 1_000.0

        self.assertAlmostEqual(
            result["monthly_crop_irrigation_l"], expected_crop_irrigation_l
        )
        self.assertAlmostEqual(result["monthly_drainage_l"], expected_drainage_l)
        self.assertAlmostEqual(result["monthly_total_water_m3"], expected_total_m3)
        self.assertAlmostEqual(
            result["monthly_water_cost_krw"], expected_total_m3 * 2_300.0
        )

    def test_drainage_ratio_must_be_less_than_one(self) -> None:
        with self.assertRaisesRegex(ValueError, "배액률"):
            calculate_water_cost(
                {"total_area_m2": 100.0},
                {"monthly_evapotranspiration_kg": 1_000.0},
                {
                    "drainage_ratio": 1.0,
                    "other_water_l_m2_day": 0.2,
                    "water_rate_krw_m3": 2_300.0,
                },
            )

    def test_all_spaces_are_combined_with_all_crops(self) -> None:
        spaces = read_csv_rows("space_info.csv")
        crops = read_csv_rows("crop_production_info.csv")
        scenarios = calculate_all_sites()

        self.assertEqual(len(scenarios), len(spaces) * len(crops))
        by_site: dict[str, set[str]] = {}
        for scenario in scenarios:
            meta = scenario["space_row"]
            self.assertIsInstance(meta, dict)
            by_site.setdefault(str(meta["site_id"]), set()).add(
                str(meta["crop_name"])
            )

        expected_site_ids = {row["site_id"] for row in spaces}
        expected_crop_names = {row["crop_name"] for row in crops}
        self.assertEqual(set(by_site), expected_site_ids)
        for site_crops in by_site.values():
            self.assertEqual(site_crops, expected_crop_names)

    def test_module_layers_are_applied_by_crop(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        space_header = (
            project_dir / "data" / "space_info.csv"
        ).read_text(encoding="utf-8-sig").splitlines()[0].split(",")
        crop_header = (
            project_dir / "data" / "crop_production_info.csv"
        ).read_text(encoding="utf-8-sig").splitlines()[0].split(",")

        self.assertNotIn("module_layers", space_header)
        self.assertIn("module_layers", crop_header)

        expected_layers = {
            row["crop_name"]: float(row["module_layers"])
            for row in read_csv_rows("crop_production_info.csv")
        }
        for scenario in calculate_all_sites():
            meta = scenario["space_row"]
            space = scenario["space"]
            self.assertIsInstance(meta, dict)
            self.assertIsInstance(space, dict)
            crop_name = str(meta["crop_name"])
            expected_layer = expected_layers[crop_name]
            expected_area = (
                float(meta["total_area_m2"])
                * float(meta["cultivable_ratio"])
                * expected_layer
            )
            self.assertEqual(float(space["module_layers"]), expected_layer)
            self.assertAlmostEqual(
                float(space["cultivation_area_m2"]), expected_area
            )

    def test_all_spaces_use_sixty_five_percent_cultivable_ratio(self) -> None:
        for space in read_csv_rows("space_info.csv"):
            self.assertEqual(float(space["cultivable_ratio"]), 0.65)

    def test_seedling_cost_uses_four_purchase_annual_equivalent(self) -> None:
        result = calculate_material_cost(
            {"cultivation_area_m2": 30.0},
            {
                "cycles_per_month": 10.0,
                "seedling_cost_per_m2_month_krw": 5_000.0,
                "daily_evapotranspiration_mm": 2.0,
            },
            {"nutrient_cost_per_l_krw": 20.0},
        )

        expected_nutrient_solution_l = 30.0 * 2.0 * 30.0 * 1.1
        expected_nutrient_cost = expected_nutrient_solution_l * 20.0
        self.assertEqual(result["monthly_seedling_cost_krw"], 150_000.0)
        self.assertAlmostEqual(
            result["monthly_nutrient_solution_l"],
            expected_nutrient_solution_l,
        )
        self.assertAlmostEqual(
            result["monthly_nutrient_cost_krw"], expected_nutrient_cost
        )
        self.assertAlmostEqual(
            result["monthly_material_cost_krw"],
            150_000.0 + expected_nutrient_cost,
        )
        self.assertEqual(result["monthly_seedling_cost_krw"] * 12, 1_800_000.0)
        self.assertEqual(30.0 * 15_000.0 * 4, 1_800_000.0)

    def test_seedling_monthly_rates_are_one_third_of_previous_prices(self) -> None:
        expected_rates = {
            "상추": 5_000.0,
            "딸기": 3_750.0,
            "바질": 5_000.0,
            "애플민트": 2_000.0,
            "쪽파": 700.0,
            "병풀": 5_000.0 / 3.0,
        }
        for crop in read_csv_rows("crop_production_info.csv"):
            self.assertAlmostEqual(
                float(crop["seedling_cost_per_m2_month_krw"]),
                expected_rates[crop["crop_name"]],
                places=2,
            )

    def test_equal_expected_income_recommends_long_term(self) -> None:
        result = calculate_profit(
            monthly_revenue_krw=1_000,
            monthly_electricity_cost_krw=0,
            monthly_water_cost_krw=0,
            monthly_material_cost_krw=0,
            monthly_labor_cost_krw=0,
            available_floor_area_m2=0,
            desired_monthly_rent_krw=800,
            standard={
                "other_cost_krw_month": 0,
                "equipment_rental_cost_krw_m2_month": 20_000,
            },
            contract={"landlord_share_ratio": 0.8},
        )

        self.assertEqual(result["landlord_expected_income_krw"], 800)
        self.assertEqual(result["recommendation"], LONG_TERM_RECOMMENDATION)
        self.assertEqual(result["contract_type"], "장기계약형")

    def test_operating_loss_is_preserved_and_recommends_short_term(self) -> None:
        result = calculate_profit(
            monthly_revenue_krw=100,
            monthly_electricity_cost_krw=200,
            monthly_water_cost_krw=0,
            monthly_material_cost_krw=0,
            monthly_labor_cost_krw=0,
            available_floor_area_m2=0,
            desired_monthly_rent_krw=0,
            standard={
                "other_cost_krw_month": 0,
                "equipment_rental_cost_krw_m2_month": 20_000,
            },
            contract={"landlord_share_ratio": 0.8},
        )

        self.assertEqual(result["monthly_operating_profit_krw"], -100)
        self.assertEqual(result["landlord_expected_income_krw"], -80)
        self.assertEqual(result["business_operating_profit_krw"], -20)
        self.assertEqual(result["recommendation"], SHORT_TERM_RECOMMENDATION)
        self.assertEqual(result["contract_type"], "단기계약형")

    def test_equipment_rental_cost_uses_available_floor_area(self) -> None:
        result = calculate_profit(
            monthly_revenue_krw=0,
            monthly_electricity_cost_krw=10_000,
            monthly_water_cost_krw=20_000,
            monthly_material_cost_krw=30_000,
            monthly_labor_cost_krw=40_000,
            available_floor_area_m2=10,
            desired_monthly_rent_krw=0,
            standard={
                "equipment_rental_cost_krw_m2_month": 20_000,
                "other_cost_krw_month": 300_000,
            },
            contract={"landlord_share_ratio": 0.8},
        )

        self.assertEqual(result["monthly_equipment_rental_cost_krw"], 200_000)
        self.assertEqual(result["monthly_base_cost_krw"], 260_000)
        self.assertEqual(result["monthly_other_cost_krw"], 300_000)
        self.assertEqual(result["monthly_operating_cost_krw"], 600_000)

    def test_excel_contains_all_scenarios_and_required_sheets(self) -> None:
        spaces = read_csv_rows("space_info.csv")
        crops = read_csv_rows("crop_production_info.csv")
        expected_count = len(spaces) * len(crops)
        scenarios = calculate_all_sites()
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "Profit_Output.xlsx"
            write_profit_output(scenarios, output_path)
            workbook = load_workbook(output_path, data_only=False)

            self.assertEqual(
                workbook.sheetnames,
                ["요약", "비용구성", "계산상세", "월별전력량", "입력기준", "검증"],
            )
            self.assertEqual(workbook["요약"].max_row, 4 + expected_count)
            self.assertEqual(workbook["요약"].max_column, 31)
            self.assertEqual(
                workbook["요약"]["A1"].value,
                f"Profit Calculator 1.0.1 · {len(spaces)}×{len(crops)} 수익성 비교",
            )
            self.assertEqual(
                workbook["요약"]["A5"].value,
                f"{spaces[0]['site_id']}-{crops[0]['crop_name']}",
            )
            self.assertEqual(
                workbook["요약"].cell(4 + expected_count, 1).value,
                f"{spaces[-1]['site_id']}-{crops[-1]['crop_name']}",
            )
            self.assertEqual(workbook["요약"]["F4"].value, "사용가능 바닥면적(m²)")
            self.assertEqual(workbook["요약"]["N4"].value, "배액률")
            self.assertEqual(workbook["요약"]["P4"].value, "월 배액량(m³)")
            self.assertEqual(workbook["요약"]["V4"].value, "월 기기 대여비(원)")
            self.assertEqual(workbook["비용구성"].max_row, 4 + expected_count)
            self.assertEqual(workbook["비용구성"].max_column, 17)
            self.assertEqual(workbook["비용구성"]["M4"].value, "기기 대여비(원)")
            self.assertEqual(workbook["비용구성"]["N4"].value, "기기 대여비 비율")
            self.assertEqual(workbook["비용구성"]["Q5"].value, "=SUM(F5,H5,J5,L5,N5,P5)")
            self.assertEqual(
                workbook["계산상세"]["E6"].value,
                "작물별 모듈 층 수",
            )
            self.assertEqual(workbook["계산상세"]["F6"].value, 4.0)
            self.assertEqual(
                workbook["월별전력량"].max_row,
                3 + expected_count * 12,
            )
            self.assertEqual(workbook["입력기준"]["C4"].value, "1.0.1")
            input_criteria = {
                workbook["입력기준"].cell(row, 2).value:
                workbook["입력기준"].cell(row, 3).value
                for row in range(4, workbook["입력기준"].max_row + 1)
            }
            self.assertEqual(input_criteria["재배가능 비율"], "0.65")
            self.assertEqual(
                input_criteria["모종비 적용 기준"],
                "기존 1회 단가의 1/3을 매월 적용",
            )
            self.assertEqual(input_criteria["양액 단가"], "20 원/L")
            self.assertEqual(
                input_criteria["작물별 모듈 층 수"],
                " · ".join(
                    f"{crop['crop_name']} {float(crop['module_layers']):g}층"
                    for crop in crops
                ),
            )
            self.assertEqual(workbook["검증"].max_row, 4 + expected_count * 3)


if __name__ == "__main__":
    unittest.main()
